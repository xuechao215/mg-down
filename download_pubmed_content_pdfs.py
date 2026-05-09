#!/usr/bin/env python3
"""Generate per-article PDFs from PubMed content referenced in data.xlsx."""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
import time
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

from lxml import etree
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "data.xlsx"
DEFAULT_OUTPUT_DIR = BASE_DIR / "pubmed_pdfs"
DEFAULT_MANIFEST = "download_manifest.json"

EFETCH_URL = (
    "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    "?db=pubmed&id={pmid}&retmode=xml"
)
REQUEST_TIMEOUT = 60
REQUEST_PAUSE_SECONDS = 0.4
MAX_FILENAME_LEN = 180
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
)

FONT_PATH = Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf")
FONT_NAME = "ArialUnicode"
PMID_RE = re.compile(r"(?<!\d)(\d{6,9})(?!\d)")


@dataclass
class WorkbookRecord:
    pmid: str
    title: str
    pubmed_url: str
    citation: str = ""
    authors: str = ""
    journal: str = ""
    publication_year: str = ""
    pmcid: str = ""
    nihms_id: str = ""
    doi: str = ""
    sheets: list[str] = field(default_factory=list)


@dataclass
class ParsedPubMed:
    article_title: str = ""
    authors: list[str] = field(default_factory=list)
    abstract_sections: list[tuple[str, str]] = field(default_factory=list)
    affiliations: list[str] = field(default_factory=list)
    publication_types: list[str] = field(default_factory=list)
    keywords: list[str] = field(default_factory=list)
    mesh_terms: list[str] = field(default_factory=list)
    grants: list[str] = field(default_factory=list)
    history: list[str] = field(default_factory=list)
    language: str = ""
    journal_title: str = ""
    journal_abbr: str = ""
    publication_date: str = ""
    doi: str = ""
    pii: str = ""
    pmcid: str = ""
    nihms: str = ""
    copyright: str = ""
    conflict_of_interest: str = ""


@dataclass
class DownloadResult:
    pmid: str
    title: str
    filename: str
    path: str
    status: str
    sheets: list[str] = field(default_factory=list)
    detail: str = ""


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value: Any) -> str:
    text = clean_text(value).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def safe_filename(name: str) -> str:
    keepchars = (" ", "-", "_", ".", "(", ")")
    cleaned = "".join(c if c.isalnum() or c in keepchars else "_" for c in clean_text(name))
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._")
    cleaned = cleaned[:MAX_FILENAME_LEN].strip(" ._")
    return cleaned or "untitled"


def column_index(headers: list[Any], *names: str) -> int | None:
    lookup = {
        normalize_header(header): idx + 1
        for idx, header in enumerate(headers)
        if clean_text(header)
    }
    for name in names:
        idx = lookup.get(normalize_header(name))
        if idx is not None:
            return idx
    return None


def extract_pmid(*values: Any) -> str:
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        if text.isdigit():
            return text
        match = PMID_RE.search(text)
        if match:
            return match.group(1)
    return ""


def load_records(workbook_path: Path) -> list[WorkbookRecord]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    deduped: dict[str, WorkbookRecord] = {}

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue

        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        title_col = column_index(headers, "Title")
        pubmed_col = column_index(headers, "pubmed_url", "PubMed URL")
        pmid_col = column_index(headers, "PMID")
        if not title_col or (not pubmed_col and not pmid_col):
            continue

        citation_col = column_index(headers, "Citation")
        authors_col = column_index(headers, "Authors")
        journal_col = column_index(headers, "Journal/Book", "Journal Book")
        year_col = column_index(headers, "Publication Year")
        pmcid_col = column_index(headers, "PMCID")
        nihms_col = column_index(headers, "NIHMS ID", "NIHMSID")
        doi_col = column_index(headers, "DOI")

        for row_idx in range(2, ws.max_row + 1):
            title = clean_text(ws.cell(row_idx, title_col).value)
            if not title:
                continue

            pubmed_url = clean_text(ws.cell(row_idx, pubmed_col).value) if pubmed_col else ""
            pmid = extract_pmid(
                ws.cell(row_idx, pmid_col).value if pmid_col else "",
                pubmed_url,
            )
            if not pmid:
                continue

            record = deduped.get(pmid)
            if record is None:
                record = WorkbookRecord(
                    pmid=pmid,
                    title=title,
                    pubmed_url=pubmed_url or f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/",
                    citation=clean_text(ws.cell(row_idx, citation_col).value) if citation_col else "",
                    authors=clean_text(ws.cell(row_idx, authors_col).value) if authors_col else "",
                    journal=clean_text(ws.cell(row_idx, journal_col).value) if journal_col else "",
                    publication_year=clean_text(ws.cell(row_idx, year_col).value) if year_col else "",
                    pmcid=clean_text(ws.cell(row_idx, pmcid_col).value) if pmcid_col else "",
                    nihms_id=clean_text(ws.cell(row_idx, nihms_col).value) if nihms_col else "",
                    doi=clean_text(ws.cell(row_idx, doi_col).value) if doi_col else "",
                    sheets=[ws.title],
                )
                deduped[pmid] = record
                continue

            if ws.title not in record.sheets:
                record.sheets.append(ws.title)

            if not record.pubmed_url and pubmed_url:
                record.pubmed_url = pubmed_url
            if not record.citation and citation_col:
                record.citation = clean_text(ws.cell(row_idx, citation_col).value)
            if not record.authors and authors_col:
                record.authors = clean_text(ws.cell(row_idx, authors_col).value)
            if not record.journal and journal_col:
                record.journal = clean_text(ws.cell(row_idx, journal_col).value)
            if not record.publication_year and year_col:
                record.publication_year = clean_text(ws.cell(row_idx, year_col).value)
            if not record.pmcid and pmcid_col:
                record.pmcid = clean_text(ws.cell(row_idx, pmcid_col).value)
            if not record.nihms_id and nihms_col:
                record.nihms_id = clean_text(ws.cell(row_idx, nihms_col).value)
            if not record.doi and doi_col:
                record.doi = clean_text(ws.cell(row_idx, doi_col).value)

    return list(deduped.values())


def fetch_pubmed_xml(pmid: str) -> str:
    url = EFETCH_URL.format(pmid=quote(pmid))
    request = Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "application/xml,text/xml;q=0.9,*/*;q=0.8",
        },
    )
    with urlopen(request, timeout=REQUEST_TIMEOUT) as response:
        charset = response.headers.get_content_charset("utf-8")
        return response.read().decode(charset, "replace")


def parse_xml(xml_text: str) -> etree._Element:
    parser = etree.XMLParser(
        recover=True,
        resolve_entities=False,
        load_dtd=False,
        no_network=True,
        remove_blank_text=True,
    )
    return etree.fromstring(xml_text.encode("utf-8"), parser=parser)


def element_text(node: etree._Element | None) -> str:
    if node is None:
        return ""
    return clean_text("".join(node.itertext()))


def build_date_text(node: etree._Element | None) -> str:
    if node is None:
        return ""
    medline = element_text(node.find("./MedlineDate"))
    if medline:
        return medline
    year = element_text(node.find("./Year"))
    month = element_text(node.find("./Month"))
    day = element_text(node.find("./Day"))
    return " ".join(part for part in (year, month, day) if part)


def parse_pubmed(xml_text: str) -> ParsedPubMed:
    root = parse_xml(xml_text)
    if root.tag != "PubmedArticleSet":
        raise ValueError(f"unexpected XML root: {root.tag}")

    article = root.find(".//PubmedArticle")
    if article is None:
        raise ValueError("no PubmedArticle node found")

    parsed = ParsedPubMed()
    parsed.article_title = element_text(article.find(".//ArticleTitle"))
    parsed.language = ", ".join(
        item for item in (element_text(node) for node in article.findall(".//Language")) if item
    )
    parsed.journal_title = element_text(article.find(".//Journal/Title"))
    parsed.journal_abbr = element_text(article.find(".//Journal/ISOAbbreviation"))
    parsed.publication_date = build_date_text(article.find(".//Journal/JournalIssue/PubDate"))

    for author in article.findall(".//AuthorList/Author"):
        collective = element_text(author.find("./CollectiveName"))
        if collective:
            author_name = collective
        else:
            fore_name = element_text(author.find("./ForeName"))
            last_name = element_text(author.find("./LastName"))
            initials = element_text(author.find("./Initials"))
            author_name = " ".join(part for part in (fore_name, last_name) if part) or initials
        if author_name:
            parsed.authors.append(author_name)

        for affiliation in author.findall("./AffiliationInfo/Affiliation"):
            text = element_text(affiliation)
            if text and text not in parsed.affiliations:
                parsed.affiliations.append(text)

    for node in article.findall(".//Abstract/AbstractText"):
        label = clean_text(node.get("Label") or node.get("NlmCategory") or "")
        text = element_text(node)
        if text:
            parsed.abstract_sections.append((label, text))

    for node in article.findall(".//OtherAbstract/AbstractText"):
        label = clean_text(node.get("Label") or node.getparent().get("Type") or "Other Abstract")
        text = element_text(node)
        if text:
            parsed.abstract_sections.append((label, text))

    for node in article.findall(".//PublicationTypeList/PublicationType"):
        text = element_text(node)
        if text and text not in parsed.publication_types:
            parsed.publication_types.append(text)

    for node in article.findall(".//KeywordList/Keyword"):
        text = element_text(node)
        if text and text not in parsed.keywords:
            parsed.keywords.append(text)

    for mesh in article.findall(".//MeshHeadingList/MeshHeading"):
        descriptor = element_text(mesh.find("./DescriptorName"))
        qualifiers = [element_text(item) for item in mesh.findall("./QualifierName") if element_text(item)]
        if descriptor:
            mesh_text = f"{descriptor} ({'; '.join(qualifiers)})" if qualifiers else descriptor
            parsed.mesh_terms.append(mesh_text)

    for grant in article.findall(".//GrantList/Grant"):
        parts = [
            element_text(grant.find("./GrantID")),
            element_text(grant.find("./Acronym")),
            element_text(grant.find("./Agency")),
            element_text(grant.find("./Country")),
        ]
        joined = " | ".join(part for part in parts if part)
        if joined:
            parsed.grants.append(joined)

    parsed.copyright = element_text(article.find(".//Abstract/CopyrightInformation"))
    parsed.conflict_of_interest = element_text(article.find(".//CoiStatement"))

    for history in article.findall(".//PubmedData/History/PubMedPubDate"):
        status = clean_text(history.get("PubStatus"))
        date_text = build_date_text(history)
        if date_text:
            line = f"{status}: {date_text}" if status else date_text
            parsed.history.append(line)

    for article_id in article.findall(".//PubmedData/ArticleIdList/ArticleId"):
        id_type = clean_text(article_id.get("IdType")).lower()
        value = element_text(article_id)
        if not value:
            continue
        if id_type == "doi":
            parsed.doi = value
        elif id_type == "pii":
            parsed.pii = value
        elif id_type == "pmc":
            parsed.pmcid = value
        elif id_type == "nihms":
            parsed.nihms = value

    return parsed


def register_fonts() -> str:
    if FONT_PATH.exists():
        if FONT_NAME not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(FONT_NAME, str(FONT_PATH)))
        return FONT_NAME
    return "Helvetica"


def build_styles() -> dict[str, ParagraphStyle]:
    font_name = register_fonts()
    sample = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "PubMedTitle",
            parent=sample["Title"],
            fontName=font_name,
            fontSize=17,
            leading=22,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#10253F"),
            spaceAfter=12,
        ),
        "meta": ParagraphStyle(
            "PubMedMeta",
            parent=sample["BodyText"],
            fontName=font_name,
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#4D6278"),
            spaceAfter=10,
        ),
        "note": ParagraphStyle(
            "PubMedNote",
            parent=sample["BodyText"],
            fontName=font_name,
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#24415F"),
            backColor=colors.HexColor("#EEF4FB"),
            borderPadding=8,
            spaceAfter=12,
        ),
        "h1": ParagraphStyle(
            "PubMedH1",
            parent=sample["Heading1"],
            fontName=font_name,
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#173A59"),
            spaceBefore=8,
            spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "PubMedBody",
            parent=sample["BodyText"],
            fontName=font_name,
            fontSize=10,
            leading=14,
            textColor=colors.black,
            spaceAfter=6,
        ),
        "bullet": ParagraphStyle(
            "PubMedBullet",
            parent=sample["BodyText"],
            fontName=font_name,
            fontSize=10,
            leading=14,
            textColor=colors.black,
            leftIndent=14,
            firstLineIndent=-10,
            spaceAfter=4,
        ),
    }


def add_section(story: list[Any], styles: dict[str, ParagraphStyle], heading: str, paragraphs: list[str]) -> None:
    items = [item for item in paragraphs if clean_text(item)]
    if not items:
        return
    story.append(Paragraph(html.escape(heading), styles["h1"]))
    for paragraph in items:
        story.append(Paragraph(html.escape(paragraph), styles["body"]))


def add_bullets(story: list[Any], styles: dict[str, ParagraphStyle], heading: str, items: list[str]) -> None:
    values = [item for item in items if clean_text(item)]
    if not values:
        return
    story.append(Paragraph(html.escape(heading), styles["h1"]))
    for item in values:
        story.append(Paragraph(f"&#8226; {html.escape(item)}", styles["bullet"]))


def meta_lines(record: WorkbookRecord, parsed: ParsedPubMed | None) -> list[str]:
    doi = (parsed.doi if parsed else "") or record.doi
    pmcid = (parsed.pmcid if parsed else "") or record.pmcid
    nihms = (parsed.nihms if parsed else "") or record.nihms_id
    journal = ", ".join(
        item
        for item in (
            parsed.journal_abbr if parsed else "",
            parsed.publication_date if parsed else "",
        )
        if item
    ) or ", ".join(item for item in (record.journal, record.publication_year) if item)

    lines = [f"PMID: {record.pmid}"]
    if journal:
        lines.append(f"Journal: {journal}")
    if doi:
        lines.append(f"DOI: {doi}")
    if pmcid:
        lines.append(f"PMCID: {pmcid}")
    if nihms:
        lines.append(f"NIHMS ID: {nihms}")
    if record.pubmed_url:
        lines.append(f"PubMed URL: {record.pubmed_url}")
    if record.sheets:
        lines.append(f"Source sheet(s): {', '.join(record.sheets)}")
    return lines


def write_pdf(
    output_path: Path,
    record: WorkbookRecord,
    parsed: ParsedPubMed | None,
    *,
    error_message: str = "",
) -> None:
    styles = build_styles()
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=42,
        rightMargin=42,
        topMargin=44,
        bottomMargin=44,
        title=record.title,
        author="Codex PubMed downloader",
    )

    title = record.title or (parsed.article_title if parsed else "") or f"PMID {record.pmid}"
    story: list[Any] = [
        Paragraph(html.escape(title), styles["title"]),
        Paragraph("<br/>".join(html.escape(line) for line in meta_lines(record, parsed)), styles["meta"]),
    ]

    if error_message:
        story.append(
            Paragraph(
                html.escape(
                    "This PDF was created with workbook metadata only because PubMed content "
                    f"could not be fetched successfully in this run. Error: {error_message}"
                ),
                styles["note"],
            )
        )
    else:
        story.append(
            Paragraph(
                html.escape("Generated from the main PubMed record content referenced by the workbook."),
                styles["note"],
            )
        )

    if record.citation:
        add_section(story, styles, "Citation", [record.citation])

    authors = parsed.authors if parsed and parsed.authors else ([record.authors] if record.authors else [])
    if authors:
        authors_text = ", ".join(authors) if isinstance(authors, list) else clean_text(authors)
        add_section(story, styles, "Authors", [authors_text])

    if parsed and parsed.abstract_sections:
        story.append(Paragraph("Abstract", styles["h1"]))
        for label, text in parsed.abstract_sections:
            if label:
                story.append(
                    Paragraph(f"<b>{html.escape(label)}.</b> {html.escape(text)}", styles["body"])
                )
            else:
                story.append(Paragraph(html.escape(text), styles["body"]))
    elif not error_message:
        story.append(Paragraph("Abstract", styles["h1"]))
        story.append(
            Paragraph(
                html.escape("No abstract text was available from the PubMed record."),
                styles["body"],
            )
        )

    if parsed:
        add_bullets(story, styles, "Publication Types", parsed.publication_types)
        add_bullets(story, styles, "Keywords", parsed.keywords)
        add_bullets(story, styles, "MeSH Terms", parsed.mesh_terms)
        add_bullets(story, styles, "Affiliations", parsed.affiliations)
        add_bullets(story, styles, "Grant Support", parsed.grants)
        add_bullets(story, styles, "PubMed Timeline", parsed.history)

        extra_sections = []
        if parsed.language:
            extra_sections.append(f"Language: {parsed.language}")
        if parsed.pii:
            extra_sections.append(f"PII: {parsed.pii}")
        if parsed.copyright:
            extra_sections.append(f"Copyright: {parsed.copyright}")
        add_section(story, styles, "Record Details", extra_sections)

        if parsed.conflict_of_interest:
            add_section(story, styles, "Conflict of Interest", [parsed.conflict_of_interest])

    story.append(Spacer(1, 6))
    doc.build(story)


def prepare_output_names(records: list[WorkbookRecord]) -> dict[str, str]:
    safe_names = {record.pmid: safe_filename(record.title) for record in records}
    counts = Counter(safe_names.values())
    filenames: dict[str, str] = {}
    for record in records:
        base = safe_names[record.pmid]
        if counts[base] > 1:
            filenames[record.pmid] = f"{base} ({record.pmid}).pdf"
        else:
            filenames[record.pmid] = f"{base}.pdf"
    return filenames


def process_records(
    records: list[WorkbookRecord],
    output_dir: Path,
    *,
    overwrite: bool,
    limit: int | None = None,
) -> list[DownloadResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    filenames = prepare_output_names(records)
    results: list[DownloadResult] = []

    selected = records[:limit] if limit else records
    for index, record in enumerate(selected, start=1):
        filename = filenames[record.pmid]
        output_path = output_dir / filename
        if output_path.exists() and not overwrite:
            results.append(
                DownloadResult(
                    pmid=record.pmid,
                    title=record.title,
                    filename=filename,
                    path=str(output_path),
                    status="skipped",
                    sheets=record.sheets[:],
                    detail="already_exists",
                )
            )
            print(f"[{index}/{len(selected)}] SKIP {filename}")
            continue

        try:
            xml_text = fetch_pubmed_xml(record.pmid)
            parsed = parse_pubmed(xml_text)
            write_pdf(output_path, record, parsed)
            results.append(
                DownloadResult(
                    pmid=record.pmid,
                    title=record.title,
                    filename=filename,
                    path=str(output_path),
                    status="generated",
                    sheets=record.sheets[:],
                    detail="pubmed_content",
                )
            )
            print(f"[{index}/{len(selected)}] OK   {filename}")
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            write_pdf(output_path, record, None, error_message=clean_text(str(exc)))
            results.append(
                DownloadResult(
                    pmid=record.pmid,
                    title=record.title,
                    filename=filename,
                    path=str(output_path),
                    status="fallback",
                    sheets=record.sheets[:],
                    detail=clean_text(str(exc)),
                )
            )
            print(f"[{index}/{len(selected)}] WARN {filename} | {clean_text(str(exc))}")
        except Exception as exc:  # pragma: no cover - defensive safeguard
            write_pdf(output_path, record, None, error_message=clean_text(str(exc)))
            results.append(
                DownloadResult(
                    pmid=record.pmid,
                    title=record.title,
                    filename=filename,
                    path=str(output_path),
                    status="fallback",
                    sheets=record.sheets[:],
                    detail=clean_text(str(exc)),
                )
            )
            print(f"[{index}/{len(selected)}] WARN {filename} | {clean_text(str(exc))}")

        time.sleep(REQUEST_PAUSE_SECONDS)

    return results


def write_manifest(output_dir: Path, results: list[DownloadResult], input_file: Path) -> Path:
    manifest_path = output_dir / DEFAULT_MANIFEST
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_file": str(input_file),
        "output_dir": str(output_dir),
        "items": [asdict(item) for item in results],
    }
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest_path


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Path to the source workbook.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for PDFs.")
    parser.add_argument("--limit", type=int, default=0, help="Only process the first N unique PubMed rows.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing PDFs.")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    input_path: Path = args.input.resolve()
    output_dir: Path = args.output_dir.resolve()
    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 2

    records = load_records(input_path)
    if not records:
        print("No PubMed-linked rows found in workbook.", file=sys.stderr)
        return 3

    results = process_records(
        records,
        output_dir,
        overwrite=args.overwrite,
        limit=args.limit or None,
    )
    manifest = write_manifest(output_dir, results, input_path)

    generated = sum(1 for item in results if item.status == "generated")
    fallback = sum(1 for item in results if item.status == "fallback")
    skipped = sum(1 for item in results if item.status == "skipped")
    print()
    print(f"Workbook: {input_path}")
    print(f"Output dir: {output_dir}")
    print(f"Manifest: {manifest}")
    print(f"Unique PubMed items: {len(records)}")
    print(f"Processed this run: {len(results)}")
    print(f"Generated: {generated} | Fallback: {fallback} | Skipped: {skipped}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
