#!/usr/bin/env python3
"""Download JTO conference PDFs from DOI rows in ACLC/WCLC workbooks."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
import time
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
VENDOR_DIR = BASE_DIR / ".vendor"
if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))

try:
    import browser_cookie3
except Exception:
    browser_cookie3 = None

try:
    from curl_cffi import requests
except Exception as exc:  # pragma: no cover - surfaced to the user immediately
    raise SystemExit(
        "Missing dependency: curl_cffi. This script expects the vendored copy in .vendor."
    ) from exc

try:
    from playwright.sync_api import sync_playwright
except Exception:
    sync_playwright = None


USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
)
PDF_ACCEPT = "application/pdf,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
REQUEST_TIMEOUT_SECONDS = 90
DOI_RESOLVE_TIMEOUT_SECONDS = 30
ITEM_DELAY_SECONDS = 0.5
SAVE_EVERY = 25
MIN_PDF_BYTES = 1024
IMPERSONATIONS = ("chrome124", "safari184")
PII_RE = re.compile(r"/retrieve/pii/([A-Z0-9]+)", re.IGNORECASE)
BLOCK_PAGE_MARKERS = (
    "just a moment",
    "enable javascript and cookies to continue",
    "performing security verification",
    "cf-mitigated",
    "access denied",
    "temporarily unavailable",
)
COOKIE_DOMAINS = {
    "sciencedirect": ("www.sciencedirect.com", "sciencedirect.com"),
    "jto": ("www.jto.org", "jto.org"),
}
TITLE_HEADERS = ("Title", "line2_title")
DOI_HEADERS = ("URL", "DOI", "line8_doi")
BROWSER_TARGETS = (
    (
        "chrome",
        "chromium",
        Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
    ),
    (
        "firefox",
        "firefox",
        Path("/Applications/Firefox.app/Contents/MacOS/firefox"),
    ),
)
PLAYWRIGHT_LAUNCH_ARGS = (
    "--disable-gpu",
    "--disable-dev-shm-usage",
    "--no-first-run",
    "--no-default-browser-check",
)
BROWSER_NAV_TIMEOUT_MS = 60_000
BROWSER_RENDER_WAIT_MS = 4_000
BROWSER_CHALLENGE_WAIT_MS = 15_000
PERSISTENT_WARMUP_WAIT_MS = 8_000
CHROME_PROFILE_ROOT = Path.home() / "Library/Application Support/Google/Chrome"
CHROME_PROFILE_NAME = "Default"
PERSISTENT_BROWSER_ARGS = PLAYWRIGHT_LAUNCH_ARGS + ("--disable-blink-features=AutomationControlled",)
PERSISTENT_ROOT_FILES = ("Local State", "Last Version", "Variations")
PERSISTENT_PROFILE_PATHS = (
    "Preferences",
    "Secure Preferences",
    "Cookies",
    "Cookies-journal",
    "Local Storage",
    "Session Storage",
    "IndexedDB",
    "Service Worker",
    "Shared Dictionary",
    "shared_proto_db",
    "WebStorage",
    "Storage",
    "Network Persistent State",
    "TransportSecurity",
    "Reporting and NEL",
    "Reporting and NEL-journal",
)


@dataclass(frozen=True)
class Preset:
    key: str
    workbook: Path
    output_dir: Path
    manifest: Path


@dataclass(frozen=True)
class Item:
    index: int
    sheet_row: int
    title: str
    doi: str


@dataclass
class Result:
    index: int
    sheet_row: int
    title: str
    doi: str
    pii: str
    status: str
    attempts: int
    file: str
    source: str
    detail: str
    size_bytes: int


PRESETS = {
    "aclc_2025": Preset(
        key="aclc_2025",
        workbook=BASE_DIR / "ACLC_2025.xlsx",
        output_dir=BASE_DIR / "ACLC",
        manifest=BASE_DIR / "ACLC" / "real_pdf_manifest.json",
    ),
    "wclc_2025": Preset(
        key="wclc_2025",
        workbook=BASE_DIR / "WCLC_2025.xlsx",
        output_dir=BASE_DIR / "WCLC",
        manifest=BASE_DIR / "WCLC" / "real_pdf_manifest.json",
    ),
}

_COOKIE_CACHE: dict[str, tuple[list[tuple[str, object | None]], list[str]]] = {}


def copy_path(src: Path, dst: Path) -> None:
    if not src.exists():
        return
    if src.is_dir():
        shutil.copytree(src, dst, dirs_exist_ok=True)
        return
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


class BrowserFallbackDownloader:
    def __init__(self) -> None:
        self._manager = None
        self._playwright = None
        self._browsers: dict[str, object] = {}
        self._contexts: dict[str, object] = {}
        self._launch_errors: dict[str, str] = {}
        self._persistent_context = None
        self._persistent_page = None
        self._persistent_profile_root: Path | None = None
        self._persistent_ready = False

    def close(self) -> None:
        if self._persistent_page is not None:
            try:
                self._persistent_page.close()
            except Exception:
                pass
        self._persistent_page = None

        if self._persistent_context is not None:
            try:
                self._persistent_context.close()
            except Exception:
                pass
        self._persistent_context = None

        for context in self._contexts.values():
            try:
                context.close()
            except Exception:
                pass
        self._contexts.clear()

        for browser in self._browsers.values():
            try:
                browser.close()
            except Exception:
                pass
        self._browsers.clear()

        if self._playwright is not None:
            try:
                self._playwright.stop()
            except Exception:
                pass
        elif self._manager is not None and hasattr(self._manager, "stop"):
            try:
                self._manager.stop()
            except Exception:
                pass

        self._manager = None
        self._playwright = None
        self._persistent_ready = False

        if self._persistent_profile_root is not None:
            shutil.rmtree(self._persistent_profile_root, ignore_errors=True)
        self._persistent_profile_root = None

    def _ensure_playwright(self) -> bool:
        if sync_playwright is None:
            self._launch_errors["playwright"] = "playwright_unavailable"
            return False

        if self._playwright is None:
            try:
                self._manager = sync_playwright()
                self._playwright = self._manager.start()
            except Exception as exc:
                self._launch_errors["playwright"] = f"start_failed:{type(exc).__name__}"
                return False

        return True

    def _context_for(self, browser_name: str, engine_name: str, executable_path: Path):
        if browser_name in self._contexts:
            return self._contexts[browser_name]

        if not executable_path.exists():
            self._launch_errors[browser_name] = "missing_browser"
            return None

        if not self._ensure_playwright():
            return None

        browser = None
        try:
            launcher = getattr(self._playwright, engine_name)
            launch_kwargs = {
                "executable_path": str(executable_path),
                "headless": True,
            }
            if engine_name == "chromium":
                launch_kwargs["args"] = list(PLAYWRIGHT_LAUNCH_ARGS)

            browser = launcher.launch(**launch_kwargs)
            context = browser.new_context(
                ignore_https_errors=True,
                user_agent=USER_AGENT,
                locale="en-US",
                extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
            )
        except Exception as exc:
            if browser is not None:
                try:
                    browser.close()
                except Exception:
                    pass
            self._launch_errors[browser_name] = f"launch_failed:{type(exc).__name__}"
            return None

        self._browsers[browser_name] = browser
        self._contexts[browser_name] = context
        return context

    def _persistent_context_for(self):
        if self._persistent_context is not None:
            return self._persistent_context

        chrome_executable = BROWSER_TARGETS[0][2]
        if not chrome_executable.exists():
            self._launch_errors["chrome_persistent"] = "missing_browser"
            return None

        if not self._ensure_playwright():
            return None

        profile_root = Path(tempfile.mkdtemp(prefix="jto-profile-"))
        profile_dir = profile_root / CHROME_PROFILE_NAME
        profile_dir.mkdir(parents=True, exist_ok=True)

        try:
            for name in PERSISTENT_ROOT_FILES:
                copy_path(CHROME_PROFILE_ROOT / name, profile_root / name)
            source_profile_dir = CHROME_PROFILE_ROOT / CHROME_PROFILE_NAME
            for name in PERSISTENT_PROFILE_PATHS:
                copy_path(source_profile_dir / name, profile_dir / name)

            context = self._playwright.chromium.launch_persistent_context(
                user_data_dir=str(profile_root),
                executable_path=str(chrome_executable),
                headless=False,
                accept_downloads=True,
                ignore_https_errors=True,
                args=list(PERSISTENT_BROWSER_ARGS),
            )
        except Exception as exc:
            shutil.rmtree(profile_root, ignore_errors=True)
            self._launch_errors["chrome_persistent"] = f"launch_failed:{type(exc).__name__}"
            return None

        self._persistent_profile_root = profile_root
        self._persistent_context = context
        self._persistent_page = context.pages[0] if context.pages else context.new_page()
        return context

    def _warm_persistent_session(self, referer: str, *, force: bool = False) -> str:
        context = self._persistent_context_for()
        if context is None:
            return self._launch_errors.get("chrome_persistent", "persistent_context_unavailable")

        if self._persistent_ready and not force:
            return "ready"

        try:
            self._persistent_page.goto(
                referer,
                wait_until="domcontentloaded",
                timeout=BROWSER_NAV_TIMEOUT_MS,
            )
            self._persistent_page.wait_for_timeout(PERSISTENT_WARMUP_WAIT_MS)
            body_text = self._persistent_page.locator("body").inner_text(timeout=5_000).lower()
        except Exception as exc:
            self._persistent_ready = False
            return f"warm_failed:{type(exc).__name__}"

        if any(marker in body_text for marker in BLOCK_PAGE_MARKERS):
            self._persistent_ready = False
            return f"warm_blocked:{compact(body_text, 120)}"

        self._persistent_ready = True
        return "ready"

    def _request_pdf_with_persistent_context(
        self,
        url: str,
        referer: str,
    ) -> tuple[bytes | None, int, str]:
        context = self._persistent_context_for()
        if context is None:
            return None, 0, self._launch_errors.get("chrome_persistent", "persistent_context_unavailable")

        warm_detail = self._warm_persistent_session(referer)
        if warm_detail != "ready":
            return None, 0, warm_detail

        attempts = 0
        last_detail = "persistent_request_failed"
        for attempt_index in range(2):
            attempts += 1
            response = None
            body = b""
            try:
                response = context.request.get(
                    url,
                    headers={
                        "Accept": PDF_ACCEPT,
                        "Referer": referer,
                    },
                    timeout=REQUEST_TIMEOUT_SECONDS * 1000,
                    fail_on_status_code=False,
                )
                body = response.body()
            except Exception as exc:
                last_detail = f"persistent_request_failed:{type(exc).__name__}"
            else:
                if response.status == 200 and len(body) >= MIN_PDF_BYTES and is_pdf_bytes(body):
                    return body, attempts, self._response_detail(response, body)
                last_detail = self._response_detail(response, body)

            if attempt_index == 0:
                warm_detail = self._warm_persistent_session(referer, force=True)
                if warm_detail != "ready":
                    return None, attempts, f"{last_detail} | {warm_detail}"

        return None, attempts, last_detail

    def _response_detail(self, response, body: bytes = b"") -> str:
        if response is None:
            return "no_response"

        try:
            content_type = response.headers.get("content-type", "")
        except Exception:
            content_type = ""

        status = getattr(response, "status", "unknown")
        detail = f"http_{status}:{content_type}"
        if body:
            snippet = body[:200].decode("utf-8", "ignore").lower()
            snippet = compact(snippet, 100)
            if snippet:
                detail += f":{snippet}"
        return detail

    def _request_pdf(self, context, url: str, referer: str) -> tuple[bytes | None, str]:
        try:
            response = context.request.get(
                url,
                headers={
                    "Accept": PDF_ACCEPT,
                    "Referer": referer,
                },
                timeout=BROWSER_NAV_TIMEOUT_MS,
                fail_on_status_code=False,
            )
        except Exception as exc:
            return None, f"request_failed:{type(exc).__name__}"

        try:
            body = response.body()
        except Exception as exc:
            return None, f"response_body_failed:{type(exc).__name__}"

        if response.status == 200 and len(body) >= MIN_PDF_BYTES and is_pdf_bytes(body):
            return body, self._response_detail(response, body)

        return None, self._response_detail(response, body)

    def _goto_pdf(self, page, url: str) -> tuple[bytes | None, str]:
        try:
            response = page.goto(url, wait_until="domcontentloaded", timeout=BROWSER_NAV_TIMEOUT_MS)
            page.wait_for_timeout(1200)
        except Exception as exc:
            return None, f"goto_failed:{type(exc).__name__}"

        if response is None:
            return None, "goto_no_response"

        try:
            body = response.body()
        except Exception as exc:
            return None, f"goto_body_failed:{type(exc).__name__}"

        if response.status == 200 and len(body) >= MIN_PDF_BYTES and is_pdf_bytes(body):
            return body, self._response_detail(response, body)

        return None, self._response_detail(response, body)

    def _wait_for_clearance(self, page) -> None:
        page.wait_for_timeout(BROWSER_RENDER_WAIT_MS)
        try:
            page.wait_for_function(
                """
                () => {
                    const title = (document.title || '').toLowerCase();
                    const body = (document.body && document.body.innerText || '').toLowerCase();
                    return !title.includes('just a moment')
                        && !body.includes('enable javascript and cookies to continue')
                        && !body.includes('performing security verification');
                }
                """,
                timeout=BROWSER_CHALLENGE_WAIT_MS,
            )
        except Exception:
            pass
        page.wait_for_timeout(1200)

    def download_pdf(self, raw_pii: str) -> tuple[bytes | None, str, int, str]:
        attempts = 0
        errors: list[str] = []

        for candidate in pdf_candidates(raw_pii):
            if candidate["source"] != "jto":
                continue
            body, request_attempts, detail = self._request_pdf_with_persistent_context(
                candidate["url"],
                candidate["referer"],
            )
            attempts += request_attempts
            if body is not None:
                return body, "browser:chrome-persistent:jto:request", attempts, detail
            errors.append(f"chrome-persistent:{candidate['source']}:request:{detail}")

        for browser_name, engine_name, executable_path in BROWSER_TARGETS:
            context = self._context_for(browser_name, engine_name, executable_path)
            if context is None:
                errors.append(f"{browser_name}:{self._launch_errors.get(browser_name, 'launch_failed')}")
                continue

            for candidate in pdf_candidates(raw_pii):
                page = context.new_page()
                try:
                    attempts += 1
                    try:
                        page.goto(
                            candidate["referer"],
                            wait_until="domcontentloaded",
                            timeout=BROWSER_NAV_TIMEOUT_MS,
                        )
                        self._wait_for_clearance(page)
                    except Exception as exc:
                        errors.append(f"{browser_name}:{candidate['source']}:prime:{type(exc).__name__}")
                        continue

                    body, detail = self._request_pdf(context, candidate["url"], candidate["referer"])
                    if body is not None:
                        source = f"browser:{browser_name}:{candidate['source']}:request"
                        return body, source, attempts, detail
                    errors.append(f"{browser_name}:{candidate['source']}:request:{detail}")

                    body, detail = self._goto_pdf(page, candidate["url"])
                    if body is not None:
                        source = f"browser:{browser_name}:{candidate['source']}:goto"
                        return body, source, attempts, detail
                    errors.append(f"{browser_name}:{candidate['source']}:goto:{detail}")
                finally:
                    try:
                        page.close()
                    except Exception:
                        pass

        return None, "", attempts, " | ".join(errors)[:600] or "browser_download_failed"


def parse_args(argv: list[str], default_conference: str | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download ACLC/WCLC JTO PDFs by resolving DOI links to Elsevier PIIs.",
    )
    parser.add_argument(
        "conference",
        nargs="?",
        choices=sorted(PRESETS),
        default=default_conference,
        help=argparse.SUPPRESS if default_conference else "Conference preset to download.",
    )
    parser.add_argument(
        "--xlsx",
        default=None,
        help="Override the workbook path from the preset.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Override the PDF output directory from the preset.",
    )
    parser.add_argument(
        "--manifest",
        default=None,
        help="Override the manifest JSON path from the preset.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only process the first N selected rows.",
    )
    parser.add_argument(
        "--start-index",
        type=int,
        default=1,
        help="1-based row index (excluding header) to start from.",
    )
    parser.add_argument(
        "--end-index",
        type=int,
        default=None,
        help="1-based row index (excluding header) to stop at, inclusive.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=ITEM_DELAY_SECONDS,
        help=f"Delay in seconds between rows (default: {ITEM_DELAY_SECONDS}).",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Redownload even if the local PDF already exists and looks valid.",
    )
    args = parser.parse_args(argv[1:])
    if not args.conference:
        parser.error("the following arguments are required: conference")
    return args


def compact(value: str, limit: int = 180) -> str:
    text = re.sub(r"\s+", " ", value or "").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def normalize_doi(value: str) -> str:
    cleaned = (value or "").strip()
    if not cleaned:
        return ""
    if cleaned.lower().startswith("doi:"):
        cleaned = cleaned[4:].strip()
    for prefix in ("https://doi.org/", "http://doi.org/"):
        if cleaned.lower().startswith(prefix):
            suffix = cleaned[len(prefix) :].strip()
            return f"https://doi.org/{suffix}"
    if cleaned.startswith("10."):
        return f"https://doi.org/{cleaned}"
    return cleaned


def find_column(headers: list[str], candidates: tuple[str, ...]) -> int:
    normalized = {header.strip().lower(): index for index, header in enumerate(headers)}
    for candidate in candidates:
        index = normalized.get(candidate.lower())
        if index is not None:
            return index
    raise SystemExit(f"Missing expected column. Tried: {', '.join(candidates)}")


def load_items(workbook_path: Path) -> list[Item]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    row_iter = worksheet.iter_rows(values_only=True)
    headers = [str(cell or "").strip() for cell in next(row_iter, ())]
    if not headers:
        return []

    title_idx = find_column(headers, TITLE_HEADERS)
    doi_idx = find_column(headers, DOI_HEADERS)

    items: list[Item] = []
    for data_index, row in enumerate(row_iter, start=1):
        title = str(row[title_idx] or "").strip().rstrip(",").strip()
        doi = normalize_doi(str(row[doi_idx] or "").strip())
        if not doi:
            continue
        if not title:
            title = doi
        items.append(
            Item(
                index=data_index,
                sheet_row=data_index + 1,
                title=title,
                doi=doi,
            )
        )
    return items


def is_pdf_bytes(content: bytes) -> bool:
    return content[:5] == b"%PDF-"


def valid_existing_pdf(path: Path) -> bool:
    if not path.exists() or path.stat().st_size < MIN_PDF_BYTES:
        return False
    try:
        return is_pdf_bytes(path.read_bytes()[:5])
    except OSError:
        return False


def format_pii(raw_pii: str) -> str:
    if len(raw_pii) < 17:
        return raw_pii
    return f"{raw_pii[:5]}-{raw_pii[5:9]}({raw_pii[9:11]}){raw_pii[11:16]}-{raw_pii[16:]}"


def parse_pii(text: str) -> str | None:
    matches = PII_RE.findall(text or "")
    if matches:
        return matches[-1].upper()

    hidden = re.search(r'name="id"\s+value="([A-Z0-9]+)"', text or "", re.IGNORECASE)
    if hidden:
        return hidden.group(1).upper()

    script_match = re.search(r"identifierValue\s*:\s*'([A-Z0-9]+)'", text or "")
    if script_match:
        return script_match.group(1).upper()

    return None


def run_curl(command: list[str], timeout_seconds: int) -> tuple[int, str]:
    try:
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds + 5,
        )
    except FileNotFoundError as exc:
        raise SystemExit("curl is required for DOI resolution but was not found.") from exc
    except subprocess.TimeoutExpired:
        return 124, ""

    output = "\n".join(part for part in (result.stdout, result.stderr) if part)
    return result.returncode, output


def resolve_pii_from_doi(doi: str) -> tuple[str | None, str]:
    header_command = [
        "curl",
        "-I",
        "-L",
        "-sS",
        "--max-time",
        str(DOI_RESOLVE_TIMEOUT_SECONDS),
        doi,
    ]
    header_rc, header_output = run_curl(header_command, DOI_RESOLVE_TIMEOUT_SECONDS)
    pii = parse_pii(header_output)
    if pii:
        return pii, "doi_headers"

    body_command = [
        "curl",
        "-L",
        "-sS",
        "--max-time",
        str(DOI_RESOLVE_TIMEOUT_SECONDS),
        doi,
    ]
    body_rc, body_output = run_curl(body_command, DOI_RESOLVE_TIMEOUT_SECONDS)
    pii = parse_pii(body_output)
    if pii:
        return pii, "doi_body"

    detail = f"pii_unresolved:header_rc={header_rc}:body_rc={body_rc}"
    if header_output:
        detail += f":{compact(header_output)}"
    elif body_output:
        detail += f":{compact(body_output)}"
    return None, detail


def cookie_choices(source: str) -> tuple[list[tuple[str, object | None]], list[str]]:
    cached = _COOKIE_CACHE.get(source)
    if cached is not None:
        return cached

    choices: list[tuple[str, object | None]] = []
    errors: list[str] = []

    if browser_cookie3 is None:
        errors.append("browser_cookie3_unavailable")
    else:
        for domain in COOKIE_DOMAINS.get(source, ()):
            try:
                jar = browser_cookie3.chrome(domain_name=domain)
            except Exception as exc:
                errors.append(f"chrome:{domain}:{type(exc).__name__}")
                continue
            choices.append((f"chrome:{domain}", jar))
            break

    choices.append(("no_cookies", None))
    _COOKIE_CACHE[source] = (choices, errors)
    return _COOKIE_CACHE[source]


def pdf_candidates(raw_pii: str) -> list[dict[str, str]]:
    formatted = format_pii(raw_pii)
    return [
        {
            "source": "jto",
            "url": f"https://www.jto.org/action/showPdf?pii={formatted}",
            "referer": f"https://www.jto.org/article/{formatted}/fulltext",
        },
        {
            "source": "sciencedirect",
            "url": f"https://www.sciencedirect.com/science/article/pii/{raw_pii}/pdfft?isDTMRedir=true&download=true",
            "referer": f"https://www.sciencedirect.com/science/article/pii/{raw_pii}",
        },
    ]


def download_via_http(raw_pii: str) -> tuple[bytes | None, str, int, str]:
    attempts = 0
    errors: list[str] = []

    for candidate in pdf_candidates(raw_pii):
        source = candidate["source"]
        cookie_plan, cookie_errors = cookie_choices(source)
        if cookie_errors:
            errors.extend(f"{source}:{detail}" for detail in cookie_errors)

        for cookie_label, cookie_jar in cookie_plan:
            for impersonation in IMPERSONATIONS:
                attempts += 1
                try:
                    response = requests.get(
                        candidate["url"],
                        headers={
                            "Accept": PDF_ACCEPT,
                            "Referer": candidate["referer"],
                            "User-Agent": USER_AGENT,
                        },
                        cookies=cookie_jar,
                        impersonate=impersonation,
                        timeout=REQUEST_TIMEOUT_SECONDS,
                        allow_redirects=True,
                    )
                except Exception as exc:
                    errors.append(f"{source}:{cookie_label}:{impersonation}:{type(exc).__name__}")
                    continue

                content_type = (response.headers.get("content-type") or "").lower()
                body = response.content or b""
                if response.status_code == 200 and len(body) >= MIN_PDF_BYTES and is_pdf_bytes(body):
                    source_label = f"{source}:{impersonation}:{cookie_label}"
                    return body, source_label, attempts, f"http_200:{content_type}"

                detail = f"{source}:{cookie_label}:{impersonation}:http_{response.status_code}:{content_type}"
                if body:
                    snippet = body[:300].decode("utf-8", "ignore").lower()
                    if any(marker in snippet for marker in BLOCK_PAGE_MARKERS):
                        detail += ":blocked"
                errors.append(detail)
                time.sleep(0.2)

    return None, "", attempts, " | ".join(errors)[:600] or "download_failed"


def load_manifest(path: Path) -> dict[str, dict]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    return {entry["doi"]: entry for entry in data if entry.get("doi")}


def ordered_manifest(items: list[Item], results_by_doi: dict[str, dict]) -> list[dict]:
    return [results_by_doi[item.doi] for item in items if item.doi in results_by_doi]


def save_manifest(path: Path, items: list[Item], results_by_doi: dict[str, dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = ordered_manifest(items, results_by_doi)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_pdf(output_path: Path, content: bytes) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(".part")
    temp_path.write_bytes(content)
    temp_path.replace(output_path)


def build_result(
    item: Item,
    pii: str,
    status: str,
    attempts: int,
    output_path: Path,
    source: str,
    detail: str,
    size_bytes: int,
) -> Result:
    return Result(
        index=item.index,
        sheet_row=item.sheet_row,
        title=item.title,
        doi=item.doi,
        pii=pii,
        status=status,
        attempts=attempts,
        file=str(output_path),
        source=source,
        detail=detail,
        size_bytes=size_bytes,
    )


def main(argv: list[str] | None = None, default_conference: str | None = None) -> int:
    if argv is None:
        argv = sys.argv
    args = parse_args(argv, default_conference=default_conference)

    preset = PRESETS[args.conference]
    workbook_path = Path(args.xlsx) if args.xlsx else preset.workbook
    output_dir = Path(args.output_dir) if args.output_dir else preset.output_dir
    manifest_path = Path(args.manifest) if args.manifest else preset.manifest

    items = load_items(workbook_path)
    selected = [
        item
        for item in items
        if item.index >= args.start_index and (args.end_index is None or item.index <= args.end_index)
    ]
    if args.limit is not None:
        selected = selected[: args.limit]

    if not selected:
        print("No rows selected.")
        return 0

    results_by_doi = load_manifest(manifest_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    browser_downloader = BrowserFallbackDownloader()

    print(
        f"Downloading {len(selected)} items from {workbook_path.name} into {output_dir}",
        flush=True,
    )
    try:
        for processed_index, item in enumerate(selected, start=1):
            previous = results_by_doi.get(item.doi)
            raw_pii = str(previous.get("pii") or "").strip() if previous else ""

            if not raw_pii:
                raw_pii, pii_detail = resolve_pii_from_doi(item.doi)
                if not raw_pii:
                    result = build_result(
                        item=item,
                        pii="",
                        status="failed",
                        attempts=0,
                        output_path=output_dir / f"row-{item.index:04d}.pdf",
                        source="doi",
                        detail=pii_detail,
                        size_bytes=0,
                    )
                    results_by_doi[item.doi] = asdict(result)
                    print(
                        f"[{processed_index}/{len(selected)}] FAIL row={item.index} pii=unresolved detail={pii_detail}",
                        flush=True,
                    )
                    if processed_index % SAVE_EVERY == 0:
                        save_manifest(manifest_path, items, results_by_doi)
                    if args.delay > 0:
                        time.sleep(args.delay)
                    continue

            output_path = output_dir / f"PII{raw_pii}.pdf"
            if not args.overwrite and valid_existing_pdf(output_path):
                result = build_result(
                    item=item,
                    pii=raw_pii,
                    status="ok",
                    attempts=0,
                    output_path=output_path,
                    source="existing_file",
                    detail="existing_file",
                    size_bytes=output_path.stat().st_size,
                )
                results_by_doi[item.doi] = asdict(result)
                print(
                    f"[{processed_index}/{len(selected)}] OK   {raw_pii} existing size={result.size_bytes}",
                    flush=True,
                )
                if processed_index % SAVE_EVERY == 0:
                    save_manifest(manifest_path, items, results_by_doi)
                if args.delay > 0:
                    time.sleep(args.delay)
                continue

            pdf_bytes, source, attempts, detail = browser_downloader.download_pdf(raw_pii)
            if pdf_bytes is None:
                http_bytes, http_source, http_attempts, http_detail = download_via_http(raw_pii)
                attempts += http_attempts
                if http_bytes is not None:
                    pdf_bytes = http_bytes
                    source = http_source
                    detail = http_detail
                else:
                    detail = f"{detail} | {http_detail}".strip(" |")

            if pdf_bytes is not None:
                write_pdf(output_path, pdf_bytes)
                result = build_result(
                    item=item,
                    pii=raw_pii,
                    status="ok",
                    attempts=attempts,
                    output_path=output_path,
                    source=source,
                    detail=detail,
                    size_bytes=len(pdf_bytes),
                )
                print(
                    f"[{processed_index}/{len(selected)}] OK   {raw_pii} attempts={attempts} via={source} size={len(pdf_bytes)}",
                    flush=True,
                )
            else:
                result = build_result(
                    item=item,
                    pii=raw_pii,
                    status="failed",
                    attempts=attempts,
                    output_path=output_path,
                    source="download",
                    detail=detail,
                    size_bytes=0,
                )
                print(
                    f"[{processed_index}/{len(selected)}] FAIL {raw_pii} attempts={attempts} detail={compact(detail, 220)}",
                    flush=True,
                )

            results_by_doi[item.doi] = asdict(result)
            if processed_index % SAVE_EVERY == 0:
                save_manifest(manifest_path, items, results_by_doi)
            if args.delay > 0:
                time.sleep(args.delay)
    finally:
        browser_downloader.close()

    save_manifest(manifest_path, items, results_by_doi)
    ordered_results = ordered_manifest(items, results_by_doi)
    ok_count = sum(1 for entry in ordered_results if entry.get("status") == "ok")
    fail_count = sum(1 for entry in ordered_results if entry.get("status") == "failed")
    print(
        f"\nSummary ok={ok_count} failed={fail_count} manifest={manifest_path}",
        flush=True,
    )
    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
